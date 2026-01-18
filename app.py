import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import re
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Set
import warnings
warnings.filterwarnings('ignore')

# Page configuration
st.set_page_config(page_title="EPA Comparison Tool", layout="wide")

# Yellow fill for highlighting
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# CANONICAL_COLUMNS will be dynamically read from Raw/Previous file template sheet

def clean_column_name(col_name: str) -> str:
    """Clean column name: strip, remove extra whitespace, remove newlines"""
    if pd.isna(col_name):
        return ""
    col_str = str(col_name)
    col_str = col_str.replace('\n', ' ').replace('\r', ' ')
    col_str = ' '.join(col_str.split())
    col_str = col_str.strip()
    return col_str

def normalize_column_name_for_matching(col_name: str) -> str:
    """Normalize column name for matching (lowercase, remove spaces and punctuation)"""
    cleaned = clean_column_name(col_name)
    return cleaned.lower().replace(' ', '').replace('.', '').replace('#', 'number')

def parse_date_from_string(s: str) -> Optional[datetime]:
    """Parse date from string (supports formats: 8.1, 8/1, 12.5, 12/5, 08.29, 08/29)"""
    if not s:
        return None
    s = s.replace("Raw Data_EPA_", "").replace("Raw Data_", "").replace("EPA_", "").replace(".xlsx", "").strip()
    pattern = r'(\d{1,2})[./](\d{1,2})'
    match = re.search(pattern, s)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        try:
            year = datetime.now().year
            return datetime(year, month, day)
        except ValueError:
            return None
    return None

def extract_date_from_sheet_or_filename(sheet_name: str, filename: str, is_latest: bool) -> Tuple[Optional[datetime], str]:
    """Extract date from sheet name or filename"""
    if is_latest:
        # Latest: prioritize filename
        date = parse_date_from_string(filename)
        if date:
            return date, f"from latest filename {filename}"
        date = parse_date_from_string(sheet_name)
        if date:
            return date, f"from latest sheet {sheet_name}"
    else:
        # Raw: prioritize sheet name
        date = parse_date_from_string(sheet_name)
        if date:
            return date, f"from raw sheet {sheet_name}"
        date = parse_date_from_string(filename)
        if date:
            return date, f"from raw filename {filename}"
    return None, ""

def identify_column_type(col_name: str, df: pd.DataFrame) -> Optional[str]:
    """Identify column type: Number, Date, Applicant Name"""
    normalized = normalize_column_name_for_matching(col_name)
    
    # Number column
    if normalized in ['number', 'no', 'no.', '#', 'numb', 'num']:
        return "Number"
    
    # Date column (exact match)
    if normalized == 'date' or normalized == 'dat':
        return "Date"
    
    # Applicant Name column
    if ('applicant' in normalized and 'name' in normalized) or normalized == 'applicantname':
        return "Applicant Name"
    
    return None

def find_template_sheet(sheets: Dict[str, pd.DataFrame]) -> Optional[str]:
    """Find the sheet with the most columns to use as template"""
    if not sheets:
        return None
    
    best_sheet = None
    max_cols = 0
    
    for sheet_name, df in sheets.items():
        col_count = len(df.columns)
        if col_count > max_cols:
            max_cols = col_count
            best_sheet = sheet_name
    
    return best_sheet

def find_header_row(worksheet, max_rows=30) -> int:
    """Find the row with the most non-empty cells to use as header row"""
    best_row = 0
    max_non_empty = 0
    
    for row_idx in range(1, min(max_rows + 1, worksheet.max_row + 1)):
        non_empty_count = 0
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip() != "":
                non_empty_count += 1
        
        if non_empty_count > max_non_empty:
            max_non_empty = non_empty_count
            best_row = row_idx
    
    return best_row if max_non_empty > 0 else 0

def normalize_column_name_for_template(col_name: str) -> str:
    """Normalize column name (only normalize Number and Date, keep others as-is)"""
    cleaned = clean_column_name(col_name)
    normalized = normalize_column_name_for_matching(cleaned)
    
    # Only normalize Number and Date
    if normalized in ['number', 'no', 'no.', '#', 'numb', 'num']:
        return "Number"
    elif normalized == 'date' or normalized == 'dat':
        return "Date"
    else:
        # Other columns keep original (only cleaned)
        return cleaned

def extract_canonical_columns_from_template(template_df: pd.DataFrame) -> List[str]:
    """Extract canonical column order from template DataFrame (normalize Number and Date, keep others)"""
    canonical_columns = []
    
    for col in template_df.columns:
        normalized_col = normalize_column_name_for_template(col)
        canonical_columns.append(normalized_col)
    
    return canonical_columns

def create_column_mapping(df: pd.DataFrame, template_columns: List[str]) -> Dict[str, str]:
    """Create column mapping: input column name → canonical column name (based on template order)"""
    mapping = {}
    
    # Build normalized mapping for input columns
    input_normalized = {}
    for col in df.columns:
        cleaned = clean_column_name(col)
        normalized = normalize_column_name_for_matching(cleaned)
        input_normalized[normalized] = col
    
    # Find corresponding input column for each template column
    for template_col in template_columns:
        template_normalized = normalize_column_name_for_matching(template_col)
        
        # Priority: exact match (after normalization)
        if template_normalized in input_normalized:
            input_col = input_normalized[template_normalized]
            if input_col not in mapping.values():
                mapping[input_col] = template_col
            continue
        
        # Special column rules (Number and Date)
        if template_col == "Number":
            for normalized, input_col in input_normalized.items():
                if normalized in ['number', 'no', 'no.', '#', 'numb', 'num']:
                    if input_col not in mapping.values():
                        mapping[input_col] = template_col
                        break
        elif template_col == "Date":
            for normalized, input_col in input_normalized.items():
                if normalized == 'date' or normalized == 'dat':
                    if input_col not in mapping.values():
                        mapping[input_col] = template_col
                        break
        else:
            # Other columns: exact match (cleaned column name)
            for input_col in df.columns:
                cleaned_input = clean_column_name(input_col)
                if cleaned_input == template_col:
                    if input_col not in mapping.values():
                        mapping[input_col] = template_col
                    break
    
    return mapping

def normalize_dataframe(df: pd.DataFrame, template_columns: List[str], mapping: Dict[str, str]) -> pd.DataFrame:
    """Normalize DataFrame: apply mapping and ensure canonical order, fill missing columns with '-'"""
    df_norm = df.copy()
    
    # Apply mapping
    df_norm.rename(columns=mapping, inplace=True)
    
    # Ensure all template columns exist, fill missing with "-"
    for col in template_columns:
        if col not in df_norm.columns:
            df_norm[col] = "-"
    
    # Reorder by template
    df_norm = df_norm.reindex(columns=template_columns)
    
    return df_norm

def normalize_value(val):
    """Normalize value for comparison (keep datetime as datetime, don't convert to string)"""
    if pd.isna(val) or val is None:
        return None
    if isinstance(val, (datetime, pd.Timestamp)):
        if isinstance(val, pd.Timestamp):
            return val.to_pydatetime()
        return val
    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return None
        return val
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return int(val)
        return val
    return val

def are_values_different(val1, val2):
    """Check if two values are different (normalized, datetime compared as datetime)"""
    norm1 = normalize_value(val1)
    norm2 = normalize_value(val2)
    
    if norm1 is None and norm2 is None:
        return False
    if norm1 is None or norm2 is None:
        return True
    
    if isinstance(norm1, datetime) and isinstance(norm2, datetime):
        return norm1 != norm2
    
    return norm1 != norm2

def convert_to_datetime(value):
    """Convert value to datetime"""
    if pd.isna(value) or value is None:
        return None
    if isinstance(value, (datetime, pd.Timestamp)):
        return value if isinstance(value, datetime) else value.to_pydatetime()
    if isinstance(value, str):
        value = value.strip()
        if value == "" or value == "-":
            return None
        date_formats = [
            '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d',
            '%m/%d/%Y', '%m-%d-%Y', '%m/%d/%y', '%m-%d-%y'
        ]
        for fmt in date_formats:
            try:
                return datetime.strptime(value, fmt)
            except:
                continue
    return None

def identify_date_columns(df: pd.DataFrame) -> Set[str]:
    """Identify date columns in DataFrame"""
    date_columns = set()
    date_keywords = ['date', 'sent', 'response', 'estimated', 'received', 'actual']
    
    for col in df.columns:
        col_lower = str(col).lower()
        if any(keyword in col_lower for keyword in date_keywords):
            date_columns.add(col)
            continue
        
        sample_values = df[col].dropna().head(100)
        if len(sample_values) > 0:
            date_count = sum(1 for val in sample_values if isinstance(val, (datetime, pd.Timestamp)) or convert_to_datetime(val) is not None)
            if date_count > len(sample_values) * 0.5:
                date_columns.add(col)
    
    return date_columns

def format_date_for_display(dt: datetime) -> str:
    """Format date for display: 8/1, 8/15, 8/29"""
    return f"{dt.month}/{dt.day}"

def load_excel_file(uploaded_file) -> Dict[str, pd.DataFrame]:
    """Load Excel file, use openpyxl to find correct header row"""
    try:
        # Read to memory first
        uploaded_file.seek(0)
        file_bytes = uploaded_file.read()
        
        # Use openpyxl to read
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        sheets = {}
        
        for sheet_name in wb.sheetnames:
            worksheet = wb[sheet_name]
            
            # Find header row (scan rows 1-30)
            header_row = find_header_row(worksheet, max_rows=30)
            
            if header_row == 0:
                # If not found, use first row
                header_row = 1
            
            # Read header (read all columns from header_row)
            headers = []
            max_col = worksheet.max_column
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=header_row, column=col_idx)
                header_value = cell.value
                if header_value is None or str(header_value).strip() == "":
                    # Keep Unnamed columns (don't delete)
                    header_value = f"Unnamed_{col_idx}"
                else:
                    header_value = str(header_value)
                headers.append(header_value)
            
            # Read data (start from header_row + 1)
            data_rows = []
            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)
                data_rows.append(row_data)
            
            # Create DataFrame (keep all columns, including Unnamed)
            df = pd.DataFrame(data_rows, columns=headers)
            
            # Clean column names (but keep Unnamed)
            cleaned_columns = {}
            for col in df.columns:
                if col.startswith("Unnamed_"):
                    cleaned_columns[col] = col  # Keep Unnamed
                else:
                    cleaned = clean_column_name(col)
                    cleaned_columns[col] = cleaned
            
            df.rename(columns=cleaned_columns, inplace=True)
            sheets[sheet_name] = df
        
        wb.close()
        return sheets
    except Exception as e:
        st.error(f"Error reading Excel file: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return {}

def create_comparison_excel(previous_df: pd.DataFrame, latest_df: pd.DataFrame,
                           previous_date: datetime, latest_date: datetime,
                           template_columns: List[str],
                           number_col: str, name_col: str, date_col: str) -> BytesIO:
    """Create comparison Excel file"""
    # Prepare output data
    previous_output = previous_df.copy()
    latest_output = latest_df.copy()
    
    # Add Date column (display only)
    previous_output[date_col] = format_date_for_display(previous_date)
    latest_output[date_col] = format_date_for_display(latest_date)
    
    # Ensure canonical order
    previous_output = previous_output[template_columns]
    latest_output = latest_output[template_columns]
    
    # Stack data
    output_df = pd.concat([previous_output, latest_output], ignore_index=True)
    
    # Identify date columns
    date_columns = identify_date_columns(output_df)
    
    # Create Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        previous_str = format_date_for_display(previous_date).replace('/', '.')
        latest_str = format_date_for_display(latest_date).replace('/', '.')
        sheet_name = f"EPA {previous_str} {latest_str}"
        output_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # ========== Phase 1: Difference comparison and highlighting ==========
        previous_numbers = set()
        if number_col in previous_df.columns:
            previous_numbers = set(previous_df[number_col].dropna().astype(str))
        
        previous_data = {}
        if number_col in previous_df.columns:
            for idx, row in previous_df.iterrows():
                num = str(row[number_col]) if pd.notna(row[number_col]) else None
                if num:
                    previous_data[num] = {}
                    for col in previous_df.columns:
                        if col != number_col and col != date_col:
                            previous_data[num][col] = row[col]
        
        cells_to_highlight = {}
        name_col_idx = None
        
        column_mapping = {}
        for idx, col in enumerate(template_columns, start=1):
            column_mapping[col] = idx
            if col == name_col:
                name_col_idx = idx
        
        previous_row_count = len(previous_output)
        latest_start_row = previous_row_count + 2
        
        for latest_idx, latest_row in latest_df.iterrows():
            excel_row = latest_start_row + latest_idx
            
            number_val = str(latest_row[number_col]) if pd.notna(latest_row[number_col]) and number_col in latest_row else None
            
            if not number_val:
                continue
            
            is_new_number = number_val not in previous_numbers
            has_any_difference = False
            
            for col in template_columns:
                if col == number_col or col == date_col:
                    continue
                
                col_idx = column_mapping.get(col)
                if col_idx is None:
                    continue
                
                latest_val = latest_row[col] if col in latest_row else None
                
                if is_new_number:
                    cells_to_highlight[(excel_row, col_idx)] = True
                    has_any_difference = True
                elif number_val in previous_data:
                    previous_val = previous_data[number_val].get(col)
                    if are_values_different(latest_val, previous_val):
                        cells_to_highlight[(excel_row, col_idx)] = True
                        has_any_difference = True
            
            if has_any_difference and name_col_idx:
                cells_to_highlight[(excel_row, name_col_idx)] = True
        
        for (row_idx, col_idx), _ in cells_to_highlight.items():
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = YELLOW_FILL
        
        # ========== Phase 2: Date formatting ==========
        if date_columns:
            column_to_idx = {col: idx + 1 for idx, col in enumerate(template_columns)}
            
            for col_name in date_columns:
                col_idx = column_to_idx.get(col_name)
                if col_idx is None:
                    continue
                
                for row_idx in range(2, len(output_df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    original_value = cell.value
                    
                    dt_value = convert_to_datetime(original_value)
                    
                    if dt_value is not None:
                        cell.value = dt_value
                        cell.number_format = 'm/d/yy'
        
        # Set formatting
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output

def render_design_logic_page():
    """Render the Design Logic documentation page"""
    st.title("Design Logic")
    
    st.markdown("""
    This page explains the design principles and technical decisions behind the EPA Comparison Tool.
    """)
    
    st.header("Why Use a Canonical Template?")
    st.markdown("""
    We use a **canonical template** for columns to ensure:
    - **Stable header names**: Output always uses consistent column names (e.g., "Number" instead of "Numb", "No.", or "#")
    - **Fixed column order**: Columns always appear in the same order, making it easy to compare outputs across different runs
    - **Consistent structure**: Even if input files have different header formats, the output follows a predictable structure
    
    The template is automatically extracted from the Raw/Previous file (the sheet with the most columns), ensuring
    the output matches your standard EPA format.
    """)
    
    st.header("How We Detect the Template")
    st.markdown("""
    The tool automatically selects the template from the **Raw/Previous file**:
    1. Scans all sheets in the Raw/Previous file
    2. Identifies the sheet with the **most columns**
    3. Uses that sheet's header row as the canonical template
    4. Normalizes only "Number" and "Date" columns (e.g., "Numb" → "Number", "Dat" → "Date")
    5. Preserves all other column names exactly as they appear in the template
    
    This ensures the output matches your standard EPA format while being flexible enough to handle variations in input files.
    """)
    
    st.header("How We Map Input Headers")
    st.markdown("""
    The tool automatically maps input column names to the canonical template:
    
    - **Number column**: Maps variations like "Numb", "No.", "#", "NUMBER", "No" → "Number"
    - **Date column**: Maps variations like "Dat", "DATE" → "Date"
    - **Other columns**: Exact match after cleaning (removing extra spaces, newlines)
    
    If a column in the Latest file doesn't match any template column, it will be filled with "-" in the output.
    """)
    
    st.header("How We Compare Values")
    st.markdown("""
    The comparison process follows these rules:
    
    1. **Exclude Number and Date from comparison**:
       - Number is the primary key (used for alignment only)
       - Date is display-only and never triggers highlighting
    
    2. **Compare raw values**:
       - Datetime values are compared as datetime objects (not converted to strings)
       - This ensures accurate date comparisons (e.g., 7/3 vs 8/29)
       - String values are trimmed and normalized
    
    3. **Date display formatting applied after comparison**:
       - All comparison logic uses raw datetime values
       - Date formatting (m/d/yy) is applied only in the final Excel output
       - This ensures formatting doesn't affect comparison results
    
    4. **Value normalization**:
       - Empty values (None, NaN, "", " ") are treated as equivalent
       - Numbers 1 and 1.0 are treated as equivalent
       - String whitespace is trimmed before comparison
    """)
    
    st.header("Why Highlight Only Latest?")
    st.markdown("""
    We highlight changes **only on the Latest rows** to:
    
    - **Focus attention on new changes**: Users can quickly see what changed in the latest version
    - **Avoid visual clutter**: Previous rows remain clean and easy to read
    - **Clear workflow**: The output shows Previous rows first (baseline), then Latest rows (with changes highlighted)
    
    Highlighting rules:
    - If a field changed: the changed cell is yellow, and Applicant Name is also yellow (if the column exists)
    - If a record is new (exists in Latest but not in Previous): the entire Latest row (except Number and Date) is highlighted
    - Date column is never highlighted (it's display-only)
    """)
    
    st.header("Header Row Detection")
    st.markdown("""
    To ensure we capture all columns correctly:
    
    1. **Scans rows 1-30**: Uses openpyxl to scan the first 30 rows of each sheet
    2. **Selects row with most non-empty cells**: Chooses the row with the maximum number of non-empty cells as the header
    3. **Preserves Unnamed columns**: Columns without headers are kept as "Unnamed_X" (they may be valid data columns)
    4. **Robust detection**: This method handles cases where headers are not in the first row
    """)

def render_comparison_page():
    """Render the main Comparison workflow page"""
    st.title("EPA Comparison Tool")
    
    # Documentation sections at the top
    with st.expander("📖 How to Use", expanded=False):
        st.markdown("""
        **Step-by-step instructions:**
        
        1. **Upload the Raw/Previous Excel file** (older version)
           - This file serves as the baseline and template
        
        2. **Upload the Latest Excel file** (newer version)
           - This file will be compared against the Previous file
        
        3. **Confirm detected dates and detected key columns**
           - Review the Preview & Diagnostics section to verify:
             - Detected dates for each file
             - Which sheet is used as template
             - Detected key columns (Number, Date, Applicant Name)
        
        4. **Click "Generate Comparison Excel"**
           - The tool will create a stacked comparison Excel file
        
        5. **Download the generated Excel**
           - The output file contains Previous rows first, then Latest rows
           - Changes are highlighted in yellow on Latest rows only
        """)
    
    with st.expander("📄 What This Tool Produces", expanded=False):
        st.markdown("""
        **Output format:**
        
        - **Stacked comparison**: Previous rows appear first, then Latest rows below
        - **Changes highlighted only on Latest rows**: Yellow highlighting appears only on the Latest version rows
        - **Primary key is Number**: All comparisons are aligned by the Number column
        - **Date is display-only**: The Date column shows version dates (e.g., 7/3, 8/29, 12/5) but never triggers highlighting
        - **Field change highlighting**: If any field changes, the changed cell is yellow and Applicant Name is also yellow (if the column exists)
        - **New record highlighting**: If a record exists in Latest but not in Previous, the Latest row is treated as "New" and the entire row (except Number and Date) is highlighted
        
        **Excel features:**
        - Frozen header row
        - AutoFilter enabled
        - Auto column width (maximum 50 characters)
        - Date cells formatted as m/d/yy (e.g., 9/28/22)
        """)
    
    # File upload section
    st.header("File Upload")
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("1. Raw/Previous File")
        st.caption("Upload the older version Excel file (serves as baseline and template)")
        raw_file = st.file_uploader("Upload Raw/Previous File", type=['xlsx', 'xls'], key="raw", label_visibility="collapsed")
    
    with col2:
        st.subheader("2. Latest File")
        st.caption("Upload the newer version Excel file (will be compared against Previous)")
        latest_file = st.file_uploader("Upload Latest File", type=['xlsx', 'xls'], key="latest", label_visibility="collapsed")
    
    if raw_file and latest_file:
        with st.spinner("Reading files..."):
            raw_sheets = load_excel_file(raw_file)
            latest_sheets = load_excel_file(latest_file)
        
        if not raw_sheets or not latest_sheets:
            st.error("Unable to read files. Please verify the file format is correct.")
            return
        
        # Find template sheet from Raw/Previous file
        template_sheet_name = find_template_sheet(raw_sheets)
        
        if not template_sheet_name:
            st.error("Unable to find template sheet from Raw/Previous file")
            return
        
        # Get template DataFrame
        template_df = raw_sheets.get(template_sheet_name)
        
        if template_df is None or len(template_df.columns) == 0:
            st.error(f"Template sheet '{template_sheet_name}' has no columns")
            return
        
        # Extract canonical column order from template (normalize Number and Date, keep others)
        CANONICAL_COLUMNS = extract_canonical_columns_from_template(template_df)
        
        st.success("✅ Files read successfully!")
        
        # Preview & Diagnostics section
        st.header("Preview & Diagnostics")
        
        # Template information
        st.subheader("Template Information")
        st.write(f"**Raw Template Sheet:** `{template_sheet_name}`")
        st.write(f"**Number of Columns:** {len(CANONICAL_COLUMNS)}")
        
        # Display first 10 and last 10 columns
        st.write("**Column Names:**")
        if len(CANONICAL_COLUMNS) <= 20:
            for i, col in enumerate(CANONICAL_COLUMNS, 1):
                st.write(f"  {i}. `{col}`")
        else:
            st.write("**First 10 columns:**")
            for i, col in enumerate(CANONICAL_COLUMNS[:10], 1):
                st.write(f"  {i}. `{col}`")
            st.write("**Last 10 columns:**")
            for i, col in enumerate(CANONICAL_COLUMNS[-10:], len(CANONICAL_COLUMNS) - 9):
                st.write(f"  {i}. `{col}`")
        
        template_columns = CANONICAL_COLUMNS
        
        # Parse dates and normalize all sheets
        raw_sheets_normalized = []
        for sheet_name, df in raw_sheets.items():
            date, source = extract_date_from_sheet_or_filename(sheet_name, raw_file.name, False)
            if date:
                mapping = create_column_mapping(df, template_columns)
                normalized_df = normalize_dataframe(df, template_columns, mapping)
                raw_sheets_normalized.append((sheet_name, normalized_df, date, source))
        
        latest_sheets_normalized = []
        latest_sheet_info = []  # Store original Latest sheet info
        for sheet_name, df in latest_sheets.items():
            date, source = extract_date_from_sheet_or_filename(sheet_name, latest_file.name, True)
            if date:
                # Record original column count
                latest_sheet_info.append((sheet_name, len(df.columns)))
                mapping = create_column_mapping(df, template_columns)
                normalized_df = normalize_dataframe(df, template_columns, mapping)
                latest_sheets_normalized.append((sheet_name, normalized_df, date, source))
        
        if not raw_sheets_normalized or not latest_sheets_normalized:
            st.error("Unable to parse dates from any sheet")
            return
        
        # Merge all sheets
        all_raw_dfs = [df for _, df, _, _ in raw_sheets_normalized]
        all_latest_dfs = [df for _, df, _, _ in latest_sheets_normalized]
        
        previous_df = pd.concat(all_raw_dfs, ignore_index=True) if all_raw_dfs else None
        latest_df = pd.concat(all_latest_dfs, ignore_index=True) if all_latest_dfs else None
        
        if previous_df is None or latest_df is None:
            st.error("Unable to merge data")
            return
        
        # Ensure canonical order, fill missing columns with "-"
        for col in template_columns:
            if col not in previous_df.columns:
                previous_df[col] = "-"
            if col not in latest_df.columns:
                latest_df[col] = "-"
        
        previous_df = previous_df.reindex(columns=template_columns)
        latest_df = latest_df.reindex(columns=template_columns)
        
        # Get dates
        previous_date, previous_source = min([(date, source) for _, _, date, source in raw_sheets_normalized], key=lambda x: x[0])
        latest_date, latest_source = max([(date, source) for _, _, date, source in latest_sheets_normalized], key=lambda x: x[0])
        
        # Detect required columns
        number_col = None
        name_col = None
        date_col = None
        
        for col in template_columns:
            col_type = identify_column_type(col, previous_df)
            if col_type == "Number" and number_col is None:
                number_col = col
            elif col_type == "Applicant Name" and name_col is None:
                name_col = col
            elif col_type == "Date" and date_col is None:
                date_col = col
        
        # Detection results
        st.subheader("Detection Results")
        st.write("**Detected Key Columns:**")
        st.write(f"- Number: `{number_col if number_col else '❌ Not found'}`")
        st.write(f"- Date: `{date_col if date_col else '❌ Not found'}`")
        st.write(f"- Applicant Name: `{name_col if name_col else '❌ Not found'}`")
        
        st.write("**Version Information:**")
        st.write(f"- Previous: {format_date_for_display(previous_date)} ({previous_source})")
        st.write(f"- Latest: {format_date_for_display(latest_date)} ({latest_source})")
        
        # Latest sheet information
        st.write("**Latest Sheet Information:**")
        if latest_sheet_info:
            for sheet_name, col_count in latest_sheet_info:
                st.write(f"- Latest Sheet Name: `{sheet_name}`")
                st.write(f"- Latest Sheet Original Column Count: {col_count}")
        if latest_sheets_normalized:
            st.write(f"- Latest Sheet Mapped Column Count: {len(template_columns)} (aligned to template, missing columns filled with '-')")
        
        st.write(f"**Output Column Count:** {len(template_columns)} (matches template)")
        
        if not number_col or not date_col:
            st.error("❌ Required columns not found (Number or Date)")
            return
        
        # Data preview (optional toggle)
        st.subheader("Data Preview")
        show_preview = st.checkbox("Show data preview (first 10 rows)", value=False)
        
        if show_preview:
            col1, col2 = st.columns(2)
            with col1:
                st.write("**Previous (first 10 rows):**")
                st.dataframe(previous_df.head(10), use_container_width=True)
            with col2:
                st.write("**Latest (first 10 rows):**")
                st.dataframe(latest_df.head(10), use_container_width=True)
        
        # Generate button
        if st.button("🚀 Generate Comparison Excel", type="primary"):
            with st.spinner("Generating comparison file..."):
                try:
                    output = create_comparison_excel(
                        previous_df, latest_df,
                        previous_date, latest_date,
                        template_columns,
                        number_col, name_col, date_col
                    )
                    
                    previous_str = format_date_for_display(previous_date).replace('/', '.')
                    latest_str = format_date_for_display(latest_date).replace('/', '.')
                    month_abbr = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"][previous_date.month - 1]
                    filename = f"EPA {month_abbr}{previous_str},{latest_str} Comparison.xlsx"
                    
                    st.success("✅ Comparison file generated successfully!")
                    
                    st.download_button(
                        label="📥 Download Comparison Excel",
                        data=output,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"Error generating file: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())

def main():
    """Main function with sidebar navigation"""
    # Sidebar navigation
    st.sidebar.title("Navigation")
    page = st.sidebar.radio(
        "Select Page",
        ["Comparison", "Design Logic"],
        label_visibility="collapsed"
    )
    
    # Render selected page
    if page == "Comparison":
        render_comparison_page()
    elif page == "Design Logic":
        render_design_logic_page()

if __name__ == "__main__":
    main()
